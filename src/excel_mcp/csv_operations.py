"""
CSV import and export operations for Excel MCP Server.

This module provides functionality for importing CSV files into Excel workbooks
and exporting Excel worksheets as CSV files.
"""
from pathlib import Path
import csv
import os
import base64
import logging
from typing import Dict, Any, List, Optional, Union, Tuple

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import DataError, FileOperationError

logger = logging.getLogger(__name__)

def import_csv_to_excel(
    filepath: Path | str,
    csv_content: str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    create_if_missing: bool = True,
    merge_mode: str = 'replace'
) -> Dict[str, Any]:
    """
    Import CSV content into an Excel workbook.
    
    Args:
        filepath: Path to the Excel file
        csv_content: Content of the CSV file as a string
        sheet_name: Name of worksheet to import into (created if doesn't exist)
        delimiter: CSV delimiter character
        create_if_missing: Whether to create the workbook if it doesn't exist
        merge_mode: How to handle existing data ('replace', 'append', 'new_sheet')
    
    Returns:
        Dict with result information
    """
    try:
        # Convert string path to Path object if needed
        if isinstance(filepath, str):
            filepath = Path(filepath)
            
        # Check if directory exists, create if necessary
        if not filepath.parent.exists():
            filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Check if file exists, create if necessary and allowed
        if not filepath.exists():
            if create_if_missing:
                wb = Workbook()
                # Remove default worksheet
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
            else:
                raise FileOperationError(f"Excel file {filepath} does not exist and create_if_missing is False")
        else:
            wb = load_workbook(filepath)
            
        # Determine target sheet
        if sheet_name is None:
            # Use first sheet or create one if workbook is empty
            if len(wb.sheetnames) > 0:
                sheet = wb.active
                actual_sheet_name = sheet.title
            else:
                sheet = wb.create_sheet("CSV_Import")
                actual_sheet_name = "CSV_Import"
        else:
            # Use specified sheet or create it
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                actual_sheet_name = sheet_name
                
                # Handle merge mode
                if merge_mode == 'replace':
                    # Clear existing sheet content
                    sheet.delete_rows(1, sheet.max_row)
                elif merge_mode == 'new_sheet':
                    # Create a new sheet with incremented name
                    base_name = sheet_name
                    counter = 1
                    new_name = f"{base_name}_{counter}"
                    while new_name in wb.sheetnames:
                        counter += 1
                        new_name = f"{base_name}_{counter}"
                    sheet = wb.create_sheet(new_name)
                    actual_sheet_name = new_name
                # For append mode, we don't need to do anything special here
            else:
                sheet = wb.create_sheet(sheet_name)
                actual_sheet_name = sheet_name
        
        # Parse CSV content
        import io
        csv_file = io.StringIO(csv_content)
        csv_reader = csv.reader(csv_file, delimiter=delimiter)
        rows = list(csv_reader)
        
        # Determine starting row for append mode
        start_row = 1
        if merge_mode == 'append' and sheet.max_row > 0:
            start_row = sheet.max_row + 1
            
        # Write to Excel
        for i, row in enumerate(rows, start=start_row):
            for j, value in enumerate(row, start=1):
                cell = sheet.cell(row=i, column=j)
                cell.value = value
                
        # Save workbook
        wb.save(filepath)
        
        return {
            "success": True,
            "message": f"Imported CSV data to {actual_sheet_name} in {filepath.name}",
            "rows_imported": len(rows),
            "columns_imported": len(rows[0]) if rows else 0,
            "sheet_name": actual_sheet_name,
            "filepath": str(filepath)
        }
        
    except Exception as e:
        logger.error(f"Error importing CSV to Excel: {e}")
        raise FileOperationError(f"Failed to import CSV: {str(e)}")


def export_worksheet_to_csv(
    filepath: Path | str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    include_header_row: bool = True
) -> Dict[str, Any]:
    """
    Export an Excel worksheet to CSV format.
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Name of worksheet to export (uses active sheet if None)
        delimiter: CSV delimiter character
        include_header_row: Whether to treat the first row as headers
    
    Returns:
        Dict with result information including CSV content
    """
    try:
        # Convert string path to Path object if needed
        if isinstance(filepath, str):
            filepath = Path(filepath)
            
        # Check if file exists
        if not filepath.exists():
            raise FileOperationError(f"Excel file {filepath} does not exist")
            
        # Load workbook
        wb = load_workbook(filepath, read_only=True)
        
        # Determine which sheet to export
        if sheet_name is None:
            sheet = wb.active
            actual_sheet_name = sheet.title
        elif sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            actual_sheet_name = sheet_name
        else:
            raise FileOperationError(f"Sheet '{sheet_name}' not found in {filepath}")
            
        # Extract data from worksheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Skip empty rows (where all cells are None)
            if any(cell is not None for cell in row):
                data.append(row)
                
        if not data:
            return {
                "success": True,
                "message": f"Exported empty sheet {actual_sheet_name} to CSV",
                "csv_content": "",
                "rows_exported": 0,
                "sheet_name": actual_sheet_name,
            }
            
        # Convert to CSV
        import io
        output = io.StringIO()
        csv_writer = csv.writer(output, delimiter=delimiter)
        
        for row in data:
            # Convert None values to empty strings
            processed_row = ["" if cell is None else str(cell) for cell in row]
            csv_writer.writerow(processed_row)
            
        csv_content = output.getvalue()
        
        return {
            "success": True,
            "message": f"Exported {actual_sheet_name} to CSV",
            "csv_content": csv_content,
            "rows_exported": len(data),
            "sheet_name": actual_sheet_name,
        }
        
    except Exception as e:
        logger.error(f"Error exporting worksheet to CSV: {e}")
        raise FileOperationError(f"Failed to export to CSV: {str(e)}")

# Helper functions for handling CSV files with base64 encoding
def import_csv_file_base64(
    filepath: Path | str,
    csv_content_base64: str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    create_if_missing: bool = True,
    merge_mode: str = 'replace'
) -> Dict[str, Any]:
    """
    Import a base64 encoded CSV file into Excel.
    
    Args:
        filepath: Path to the Excel file
        csv_content_base64: Base64 encoded CSV content
        Other args: Same as import_csv_to_excel
    
    Returns:
        Dict with result information
    """
    try:
        # Decode base64 content
        csv_bytes = base64.b64decode(csv_content_base64)
        csv_content = csv_bytes.decode('utf-8')
        
        # Import the CSV content
        return import_csv_to_excel(
            filepath, 
            csv_content, 
            sheet_name, 
            delimiter, 
            create_if_missing,
            merge_mode
        )
    except Exception as e:
        logger.error(f"Error importing base64 CSV: {e}")
        raise FileOperationError(f"Failed to import base64 CSV: {str(e)}")
        
def export_worksheet_to_csv_base64(
    filepath: Path | str,
    sheet_name: Optional[str] = None,
    delimiter: str = ',',
    include_header_row: bool = True
) -> Dict[str, Any]:
    """
    Export an Excel worksheet to base64 encoded CSV.
    
    Args:
        Same as export_worksheet_to_csv
    
    Returns:
        Dict with result information including base64 encoded CSV content
    """
    try:
        # Export to CSV
        result = export_worksheet_to_csv(
            filepath,
            sheet_name,
            delimiter,
            include_header_row
        )
        
        # Encode to base64
        csv_bytes = result["csv_content"].encode('utf-8')
        csv_content_base64 = base64.b64encode(csv_bytes).decode('utf-8')
        
        # Update result with base64 content
        result["csv_content_base64"] = csv_content_base64
        
        return result
    except Exception as e:
        logger.error(f"Error exporting to base64 CSV: {e}")
        raise FileOperationError(f"Failed to export to base64 CSV: {str(e)}")

# New Bulk CSV Operations

def bulk_import_csv_to_excel(
    filepath: Path | str,
    csv_data_list: List[Dict[str, Any]],
    create_if_missing: bool = True,
) -> List[Dict[str, Any]]:
    """
    Import multiple CSV contents into an Excel workbook.
    
    Args:
        filepath: Path to the Excel file
        csv_data_list: List of dictionaries containing CSV information:
            - csv_content: Content of the CSV file as a string
            - sheet_name: Name of worksheet to import into
            - delimiter: CSV delimiter character
            - merge_mode: How to handle existing data
        create_if_missing: Whether to create the workbook if it doesn't exist
    
    Returns:
        List of Dict with results for each import operation
    """
    results = []
    
    try:
        # Convert string path to Path object if needed
        if isinstance(filepath, str):
            filepath = Path(filepath)
            
        # Process each CSV import
        for csv_data in csv_data_list:
            try:
                csv_content = csv_data.get("csv_content", "")
                sheet_name = csv_data.get("sheet_name")
                delimiter = csv_data.get("delimiter", ",")
                merge_mode = csv_data.get("merge_mode", "replace")
                
                # Import the CSV content
                result = import_csv_to_excel(
                    filepath,
                    csv_content,
                    sheet_name,
                    delimiter,
                    create_if_missing,
                    merge_mode
                )
                
                results.append({
                    "success": True,
                    "sheet_name": result["sheet_name"],
                    "rows_imported": result["rows_imported"],
                    "columns_imported": result.get("columns_imported", 0),
                    "message": result["message"]
                })
                
                # After first import, the file exists
                create_if_missing = False
                
            except Exception as e:
                sheet_info = f" to sheet '{sheet_name}'" if sheet_name else ""
                results.append({
                    "success": False,
                    "sheet_name": sheet_name,
                    "message": f"Failed to import CSV{sheet_info}: {str(e)}"
                })
                
        return results
        
    except Exception as e:
        logger.error(f"Error in bulk CSV import: {e}")
        raise FileOperationError(f"Failed to perform bulk CSV import: {str(e)}")

def bulk_import_csv_file_base64(
    filepath: Path | str,
    csv_data_list: List[Dict[str, Any]],
    create_if_missing: bool = True,
) -> List[Dict[str, Any]]:
    """
    Import multiple base64 encoded CSV files into an Excel workbook.
    
    Args:
        filepath: Path to the Excel file
        csv_data_list: List of dictionaries containing CSV information:
            - csv_content_base64: Base64 encoded content of the CSV file
            - sheet_name: Name of worksheet to import into
            - delimiter: CSV delimiter character
            - merge_mode: How to handle existing data
        create_if_missing: Whether to create the workbook if it doesn't exist
    
    Returns:
        List of Dict with results for each import operation
    """
    decoded_data_list = []
    
    try:
        # Decode each base64 CSV content
        for csv_data in csv_data_list:
            try:
                csv_content_base64 = csv_data.get("csv_content_base64", "")
                csv_bytes = base64.b64decode(csv_content_base64)
                csv_content = csv_bytes.decode('utf-8')
                
                decoded_data = {
                    "csv_content": csv_content,
                    "sheet_name": csv_data.get("sheet_name"),
                    "delimiter": csv_data.get("delimiter", ","),
                    "merge_mode": csv_data.get("merge_mode", "replace")
                }
                
                decoded_data_list.append(decoded_data)
                
            except Exception as e:
                sheet_name = csv_data.get("sheet_name", "unknown")
                return [{
                    "success": False,
                    "sheet_name": sheet_name,
                    "message": f"Failed to decode base64 CSV for sheet '{sheet_name}': {str(e)}"
                }]
        
        # Process the decoded CSV data
        return bulk_import_csv_to_excel(
            filepath,
            decoded_data_list,
            create_if_missing
        )
        
    except Exception as e:
        logger.error(f"Error in bulk base64 CSV import: {e}")
        raise FileOperationError(f"Failed to perform bulk base64 CSV import: {str(e)}")

def bulk_export_worksheets_to_csv(
    filepath: Path | str,
    worksheet_list: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Export multiple Excel worksheets to CSV format.
    
    Args:
        filepath: Path to the Excel file
        worksheet_list: List of dictionaries containing worksheet information:
            - sheet_name: Name of worksheet to export
            - delimiter: CSV delimiter character
            - include_header_row: Whether to treat the first row as headers
    
    Returns:
        List of Dict with results for each export operation including CSV content
    """
    results = []
    
    try:
        # Convert string path to Path object if needed
        if isinstance(filepath, str):
            filepath = Path(filepath)
            
        # Check if file exists
        if not filepath.exists():
            raise FileOperationError(f"Excel file {filepath} does not exist")
            
        # Process each worksheet export
        for worksheet_data in worksheet_list:
            try:
                sheet_name = worksheet_data.get("sheet_name")
                delimiter = worksheet_data.get("delimiter", ",")
                include_header_row = worksheet_data.get("include_header_row", True)
                
                # Export the worksheet to CSV
                result = export_worksheet_to_csv(
                    filepath,
                    sheet_name,
                    delimiter,
                    include_header_row
                )
                
                results.append({
                    "success": True,
                    "sheet_name": result["sheet_name"],
                    "rows_exported": result["rows_exported"],
                    "columns_exported": result.get("columns_exported", 0),
                    "csv_content": result["csv_content"],
                    "message": result["message"]
                })
                
            except Exception as e:
                results.append({
                    "success": False,
                    "sheet_name": sheet_name,
                    "message": f"Failed to export worksheet '{sheet_name}' to CSV: {str(e)}"
                })
                
        return results
        
    except Exception as e:
        logger.error(f"Error in bulk worksheet CSV export: {e}")
        raise FileOperationError(f"Failed to perform bulk worksheet CSV export: {str(e)}")

def bulk_export_worksheets_to_csv_base64(
    filepath: Path | str,
    worksheet_list: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Export multiple Excel worksheets to base64 encoded CSV.
    
    Args:
        filepath: Path to the Excel file
        worksheet_list: List of dictionaries containing worksheet information:
            - sheet_name: Name of worksheet to export
            - delimiter: CSV delimiter character
            - include_header_row: Whether to treat the first row as headers
    
    Returns:
        List of Dict with results for each export including base64 encoded CSV content
    """
    results = []
    
    try:
        # Export the worksheets to CSV
        csv_results = bulk_export_worksheets_to_csv(
            filepath,
            worksheet_list
        )
        
        # Base64 encode each CSV content
        for result in csv_results:
            if result["success"]:
                csv_bytes = result["csv_content"].encode('utf-8')
                csv_content_base64 = base64.b64encode(csv_bytes).decode('utf-8')
                
                # Update result with base64 content and remove plain text content
                result["csv_content_base64"] = csv_content_base64
                del result["csv_content"]
                
            results.append(result)
                
        return results
        
    except Exception as e:
        logger.error(f"Error in bulk worksheet base64 CSV export: {e}")
        raise FileOperationError(f"Failed to perform bulk worksheet base64 CSV export: {str(e)}")
